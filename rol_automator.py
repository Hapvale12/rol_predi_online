import pandas as pd
import sys
import locale
import requests
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

# Importamos re y os aquí porque los usaremos dentro de la función
import re 
import os
from feriados_logic import cargar_feriados_locales, obtener_feriados_semana

def generar_excel_desde_texto(texto_whatsapp, fecha_inicio_str, project_root_path='.'):
    """
    Función principal que encapsula toda la lógica de procesamiento y generación de Excel.
    Recibe el texto de WhatsApp, la fecha de inicio y la ruta raíz del proyecto.
    Devuelve el libro de trabajo de openpyxl y el nombre del archivo.
    """
    # --- 2. EXTRACCIÓN Y TRANSFORMACIÓN DE DATOS ---
    REGEX_PATTERN = re.compile(
        r"^(?P<Dia>[A-ZÁÉÍÓÚÑa-zñáéíóú]+)\s+"
        r"(?P<Hora>\d{1,2}:\d{2})\s*(?P<Periodo>am|pm)?\.?\s*"
        r"(?P<Cuerpo>.+)$"
    )

    # --- 3. LÓGICA DE FECHAS Y FERIADOS ---
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, "%d/%m/%Y")
        fecha_fin = fecha_inicio + pd.Timedelta(days=6)
    except ValueError:
        raise ValueError("Formato de fecha inválido. Usa DD/MM/AAAA.")

    try:
        locale.setlocale(locale.LC_TIME, 'es-ES')
    except locale.Error:
        print("⚠️ Alerta: No se pudo configurar el idioma a español.")

    def normalizar_dia(nombre_dia):
        return nombre_dia.lower().replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')

    fechas_semana = pd.date_range(start=fecha_inicio, end=fecha_fin)
    mapa_dias_a_fechas = {normalizar_dia(fecha.strftime('%A')): fecha.date() for fecha in fechas_semana}
    
    # Usar la ruta del proyecto para encontrar el JSON de feriados
    ruta_feriados_json = os.path.join(project_root_path, 'feriados.json')
    feriados_en_semana = obtener_feriados_semana(fecha_inicio, fecha_fin, ruta_json=ruta_feriados_json)

    datos_limpios = []
    for linea in texto_whatsapp.strip().split('\n'):
        linea = linea.strip()
        if not linea.lower().startswith(('lunes', 'martes', 'miercoles', 'miércoles', 'jueves', 'viernes', 'sabado', 'sábado', 'domingo')):
            continue

        match = REGEX_PATTERN.search(linea)
        if not match:
            continue

        data = match.groupdict()
        cuerpo = data['Cuerpo'].strip()
        
        conductor_pattern = r"\b((?:[A-ZÁÉÍÓÚÑ][a-zñáéíóú]+\s+){1,2}[A-ZÁÉÍÓÚÑ][a-zñáéíóú]+)\b"
        possible_conductors = [m for m in re.finditer(conductor_pattern, cuerpo) if not m.group(1).lower().startswith(('avenida', 'parque', 'calle', 'salon'))]
        conductor = possible_conductors[-1].group(1).strip() if possible_conductors else "ERROR"

        terr_pattern = r"(?:Territorio\s*)?((?:\d+\s*parte\s+[a-zA-Z](?:\.\s*[\w\s]+?)?)(?=\.\s[A-ZÁÉÍÓÚÑ]|$)|(?:\b\d[\d\s,y]*\d\b)|(?<=\s)\d+(?=\.|\s|$))"
        terr_match = re.search(terr_pattern, cuerpo, re.IGNORECASE)

        if terr_match:
            territorio_bruto = terr_match.group(1).strip()
            if conductor != "ERROR" and territorio_bruto in conductor:
                territorio_bruto = ""
                terr_final = "T.XX"
            else:
                terr_limpio = re.sub(r'(\d+)\s*parte\s*([a-zA-Z])', r'\1\2', territorio_bruto, flags=re.IGNORECASE)
                terr_limpio = re.sub(r'\s*y\s*', ',', terr_limpio, flags=re.IGNORECASE)
                terr_limpio = re.sub(r'\s*,\s*', ',', terr_limpio)
                terr_final = 'T.' + re.sub(r'\s+', '', terr_limpio).replace('.', '').upper()
                terr_final = terr_final.replace('ALASANIMAS', 'Animas')
        else:
            territorio_bruto = ""
            terr_final = "T.XX"

        lugar_bruto = cuerpo
        if conductor != "ERROR":
            lugar_bruto = lugar_bruto.replace(conductor, '')
        if terr_match and territorio_bruto:
            lugar_bruto = lugar_bruto.replace(terr_match.group(0), '')
        
        lugar_bruto = re.sub(r'[\s\.]+', ' ', lugar_bruto).strip().replace('Avenida', 'Av.')

        hora_final = data['Hora']
        periodo = data.get('Periodo')
        if periodo:
            hora_final += periodo
        else:
            if data['Dia'].lower() in ['lunes', 'martes', 'miercoles', 'miércoles', 'jueves', 'viernes'] and '7:00' in hora_final:
                hora_final += 'pm'
            elif data['Dia'].lower() not in ['sabado', 'domingo']:
                hora_final += 'am'

        dia_normalizado = normalizar_dia(data['Dia'])
        fecha_asignacion = mapa_dias_a_fechas.get(dia_normalizado)
        if hora_final == "9:00am" and fecha_asignacion in feriados_en_semana:
            hora_final = "9:30am"

        datos_limpios.append({'DÍA': data['Dia'].capitalize(), 'HORA': hora_final, 'CONDUCTOR': conductor, 'LUGAR_BRUTO': lugar_bruto, 'TERRITORIO': terr_final})

    df = pd.DataFrame(datos_limpios)
    df['LUGAR DE PREDICACION'] = df['LUGAR_BRUTO']
    df['MÉTODO DE PREDICACIÓN'] = 'PREDICACION Y NO EN CASA P.'
    df['ID_ASIGNACION'] = range(1, len(df) + 1)
    columnas_finales = ['ID_ASIGNACION', 'DÍA', 'HORA', 'CONDUCTOR', 'LUGAR DE PREDICACION', 'MÉTODO DE PREDICACIÓN', 'TERRITORIO']
    df_final = df[columnas_finales]

    HOJA_DATOS = 'Datos_Crudos'
    PLANTILLA_FILE = os.path.join(project_root_path, 'template', 'template.xlsx')
    wb = load_workbook(PLANTILLA_FILE)
    if HOJA_DATOS in wb.sheetnames:
        del wb[HOJA_DATOS]
    ws_datos = wb.create_sheet(HOJA_DATOS)
    for r_idx, row in enumerate(dataframe_to_rows(df_final, header=True, index=False)):
        ws_datos.append(row)

    HOJA_PRINCIPAL = 'Final'
    if HOJA_PRINCIPAL in wb.sheetnames:
        ws_principal = wb[HOJA_PRINCIPAL]
        ws_principal.merge_cells('A1:B2')
        
        IMG_FILE = os.path.join(project_root_path, 'template', 'img1.png')
        img = Image(IMG_FILE)
        img.anchor = 'A1'
        img.height = 100
        img.width = 170
        ws_principal.add_image(img)

    if fecha_inicio.strftime('%B') == fecha_fin.strftime('%B'):
        rango_fechas_texto = f"Del {fecha_inicio.day} al {fecha_fin.day} de {fecha_inicio.strftime('%B')}"
    else:
        rango_fechas_texto = f"Del {fecha_inicio.day} de {fecha_inicio.strftime('%B')} al {fecha_fin.day} de {fecha_fin.strftime('%B')}"

    ws_datos['A' + str(len(df_final) + 3)] = rango_fechas_texto
    output_filename = f"ROL-({rango_fechas_texto} {fecha_inicio.year}).xlsx"
    
    return wb, output_filename

if __name__ == '__main__':
    # --- 1. DATOS DE ENTRADA INTERACTIVA (SOLO SI SE EJECUTA DIRECTAMENTE) ---
    print("----------------------------------------------------------")
    print("  PROGRAMA DE ASIGNACIONES - EXTRACCIÓN DE TEXTO")
    print("----------------------------------------------------------")
    print("PASO 1: Copia el texto completo de las asignaciones de WhatsApp.")
    print("PASO 2: Pégalo aquí (puedes pegar varias líneas a la vez).")
    print("PASO 3: Una vez pegado, presiona Enter y luego Ctrl+Z (o Ctrl+D) y luego Enter de nuevo para finalizar la entrada.")
    print("----------------------------------------------------------")

    try:
        texto_whatsapp_input = sys.stdin.read()
        if not texto_whatsapp_input.strip():
            raise EOFError
    except EOFError:
        print("\nNo se detectó entrada. Por favor, pega el texto de WhatsApp manualmente:")
        texto_whatsapp_input = input("Pega aquí el texto de WhatsApp: ")

    fecha_inicio_input = input("\nIngresa la fecha de inicio de la semana (DD/MM/AAAA): ")

    try:
        print("\nIniciando proceso...")
        # Al ejecutar localmente, la raíz es el directorio actual ('.')
        workbook, filename = generar_excel_desde_texto(texto_whatsapp_input, fecha_inicio_input, '.')
        
        # Crear el directorio 'output' si no existe
        if not os.path.exists('./output'):
            os.makedirs('./output')
        output_path = f"./output/{filename}"
        
        workbook.save(output_path)
        print(f"\n✅ Proceso completado exitosamente.")
        print(f"Archivo generado: {output_path}")
    except FileNotFoundError as e:
        print(f"\n❌ ERROR: No se encontró un archivo de plantilla. Detalles: {e}")
    except Exception as e:
        print(f"\n❌ ERROR INESPERADO: {e}")